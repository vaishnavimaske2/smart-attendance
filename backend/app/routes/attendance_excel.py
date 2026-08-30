from datetime import date, timedelta
from io import BytesIO

from fastapi import (
    APIRouter,
    Depends,
    Query,
    HTTPException
)

from fastapi.responses import StreamingResponse

from sqlalchemy.orm import Session
from sqlalchemy import func, distinct

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.database.database import get_db

from app.database.models import (
    User,
    Student,
    Class,
    Subject,
    ClassSubject,
    TeacherAssignment,
    Attendance
)

from app.core.security import get_current_user


router = APIRouter(
    prefix="/api/attendance-excel",
    tags=["Attendance Excel"]
)


# ============================================================
# EXPORT ATTENDANCE EXCEL
# ============================================================

@router.get("/export")
def export_attendance_excel(

    class_name: str = Query(...),

    section: str = Query(...),

    subject_name: str = Query(...),

    period: str = Query(
        "whole",
        description=(
            "Export period: "
            "weekly, monthly, whole or custom"
        )
    ),

    from_date: date | None = Query(None),

    to_date: date | None = Query(None),

    db: Session = Depends(get_db),

    current_user: User = Depends(
        get_current_user
    )
):

    # ========================================================
    # 1. ONLY ADMIN OR TEACHER
    # ========================================================

    if current_user.role not in {"ADMIN", "TEACHER"}:

        raise HTTPException(
            status_code=403,
            detail="Only admin or teacher can export attendance"
        )


    # ========================================================
    # 2. CLEAN INPUT
    # ========================================================

    class_name = class_name.strip()

    section = section.strip().upper()

    subject_name = subject_name.strip()

    period = period.strip().lower()


    # ========================================================
    # 3. VALIDATE PERIOD
    # ========================================================

    allowed_periods = {
        "weekly",
        "monthly",
        "whole",
        "custom"
    }

    if period not in allowed_periods:

        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid export period. "
                "Use weekly, monthly, whole or custom."
            )
        )


    # ========================================================
    # 4. VALIDATE CUSTOM DATE RANGE
    # ========================================================

    if period == "custom":

        if not from_date or not to_date:

            raise HTTPException(
                status_code=400,
                detail=(
                    "From date and To date are required "
                    "for custom export."
                )
            )

        if from_date > to_date:

            raise HTTPException(
                status_code=400,
                detail=(
                    "From date cannot be after To date."
                )
            )


    # ========================================================
    # 5. FIND CLASS
    # ========================================================

    classroom = (
        db.query(Class)
        .filter(
            Class.school_id
            == current_user.school_id,

            func.lower(Class.name)
            == class_name.lower(),

            func.upper(Class.section)
            == section
        )
        .first()
    )


    if not classroom:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Class '{class_name}' "
                f"with section '{section}' not found"
            )
        )


    # ========================================================
    # 6. FIND SUBJECT
    # ========================================================

    subject = (
        db.query(Subject)
        .filter(

            Subject.school_id
            == current_user.school_id,

            func.lower(Subject.name)
            == subject_name.lower(),

            Subject.is_active == True
        )
        .first()
    )


    if not subject:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Subject '{subject_name}' not found"
            )
        )


    # ========================================================
    # 7. CHECK SUBJECT BELONGS TO CLASS
    # ========================================================

    class_subject = (
        db.query(ClassSubject)
        .filter(
            ClassSubject.class_id
            == classroom.id,

            ClassSubject.subject_id
            == subject.id,

            ClassSubject.is_active == True
        )
        .first()
    )


    if not class_subject:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Subject '{subject.name}' "
                f"is not assigned to "
                f"{classroom.name} - "
                f"Section {classroom.section}"
            )
        )


    # ========================================================
    # 8. CHECK TEACHER ASSIGNMENT
    # ========================================================

    if current_user.role == "TEACHER":

        teacher_assignment = (
            db.query(TeacherAssignment)
            .filter(
                TeacherAssignment.teacher_id
                == current_user.id,

                TeacherAssignment.class_id
                == classroom.id,

                TeacherAssignment.subject_id
                == subject.id,

                TeacherAssignment.is_active == True
            )
            .first()
        )

        if not teacher_assignment:

            raise HTTPException(
                status_code=403,
                detail=(
                    f"You are not assigned to teach "
                    f"{subject.name} in "
                    f"{classroom.name} - "
                    f"Section {classroom.section}"
                )
            )


    # ========================================================
    # 9. DETERMINE DATE RANGE
    # ========================================================

    today = date.today()

    export_from_date = None
    export_to_date = today


    # --------------------------------------------------------
    # WEEKLY
    # --------------------------------------------------------

    if period == "weekly":

        # Current week: Monday -> today

        export_from_date = (
            today
            - timedelta(
                days=today.weekday()
            )
        )


    # --------------------------------------------------------
    # MONTHLY
    # --------------------------------------------------------

    elif period == "monthly":

        export_from_date = today.replace(
            day=1
        )


    # --------------------------------------------------------
    # WHOLE ATTENDANCE
    # --------------------------------------------------------

    elif period == "whole":

        # No lower date limit.
        # Attendance records determine the actual range.

        export_from_date = None

        export_to_date = today


    # --------------------------------------------------------
    # CUSTOM
    # --------------------------------------------------------

    elif period == "custom":

        export_from_date = from_date

        export_to_date = to_date


    # ========================================================
    # 10. GET ACTIVE STUDENTS
    # ========================================================

    students = (
        db.query(Student)
        .filter(
            Student.school_id
            == current_user.school_id,

            Student.class_id
            == classroom.id,

            Student.is_active == True
        )
        .order_by(
            Student.roll_number
        )
        .all()
    )


    if not students:

        raise HTTPException(
            status_code=404,
            detail="No active students found"
        )


    # ========================================================
    # 11. BASE ATTENDANCE QUERY
    # ========================================================

    attendance_query = (
        db.query(Attendance)
        .filter(
            Attendance.school_id
            == current_user.school_id,

            Attendance.class_id
            == classroom.id,

            Attendance.subject_id
            == subject.id
        )
    )


    # ========================================================
    # 12. APPLY DATE RANGE
    # ========================================================

    if export_from_date is not None:

        attendance_query = attendance_query.filter(
            Attendance.attendance_date
            >= export_from_date
        )


    if export_to_date is not None:

        attendance_query = attendance_query.filter(
            Attendance.attendance_date
            <= export_to_date
        )


    # ========================================================
    # 13. GET ATTENDANCE RECORDS
    # ========================================================

    attendance_records = (
        attendance_query
        .order_by(
            Attendance.attendance_date
        )
        .all()
    )


    # ========================================================
    # 14. COUNT TOTAL LECTURES
    # ========================================================

    lecture_dates = {

        attendance.attendance_date

        for attendance in attendance_records

    }


    total_lectures = len(
        lecture_dates
    )


    # ========================================================
    # 15. CREATE ATTENDANCE LOOKUP
    # ========================================================

    attendance_map = {}


    for attendance in attendance_records:

        student_id = attendance.student_id

        if student_id not in attendance_map:

            attendance_map[student_id] = []

        attendance_map[
            student_id
        ].append(attendance)


    # ========================================================
    # 16. CREATE EXCEL WORKBOOK
    # ========================================================

    workbook = Workbook()

    worksheet = workbook.active

    worksheet.title = "Attendance"


    # ========================================================
    # 17. TITLE
    # ========================================================

    worksheet["A1"] = (
        "Student Attendance Report"
    )

    worksheet["A1"].font = Font(
        bold=True,
        size=16
    )

    worksheet.merge_cells(
        "A1:H1"
    )


    # ========================================================
    # 18. CLASS INFORMATION
    # ========================================================

    worksheet["A2"] = "Class"

    worksheet["B2"] = classroom.name

    worksheet["C2"] = "Section"

    worksheet["D2"] = classroom.section

    worksheet["E2"] = "Subject"

    worksheet["F2"] = subject.name

    worksheet["G2"] = "Export Period"

    worksheet["H2"] = period.title()


    # ========================================================
    # 19. DATE INFORMATION
    # ========================================================

    worksheet["A3"] = "From Date"

    worksheet["B3"] = (
        export_from_date
        if export_from_date
        else "All Records"
    )

    worksheet["C3"] = "To Date"

    worksheet["D3"] = export_to_date

    worksheet["E3"] = "Total Lectures"

    worksheet["F3"] = total_lectures


    # ========================================================
    # 20. HEADERS
    # ========================================================

    headers = [

        "Roll Number",

        "Name",

        "Division",

        "Gender",

        "Total Lectures",

        "Present",

        "Absent",

        "Attendance %"

    ]


    header_row = 5


    for column_number, header in enumerate(
        headers,
        start=1
    ):

        cell = worksheet.cell(
            row=header_row,
            column=column_number
        )

        cell.value = header

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center"
        )


    # ========================================================
    # 21. STUDENT DATA
    # ========================================================

    current_row = 6


    for student in students:

        student_records = (
            attendance_map.get(
                student.id,
                []
            )
        )


        # ----------------------------------------------------
        # PRESENT
        # ----------------------------------------------------

        present_lectures = sum(

            1

            for attendance
            in student_records

            if attendance.status.upper()
            in {
                "PRESENT",
                "LATE"
            }

        )


        # ----------------------------------------------------
        # ABSENT
        # ----------------------------------------------------

        absent_lectures = (
            total_lectures
            - present_lectures
        )


        if absent_lectures < 0:

            absent_lectures = 0


        # ----------------------------------------------------
        # PERCENTAGE
        # ----------------------------------------------------

        if total_lectures > 0:

            attendance_percentage = (
                present_lectures
                / total_lectures
            )

        else:

            attendance_percentage = 0


        # ----------------------------------------------------
        # WRITE ROW
        # ----------------------------------------------------

        worksheet.cell(
            row=current_row,
            column=1,
            value=student.roll_number
        )


        worksheet.cell(
            row=current_row,
            column=2,
            value=student.name
        )


        worksheet.cell(
            row=current_row,
            column=3,
            value=student.section
        )


        worksheet.cell(
            row=current_row,
            column=4,
            value=student.gender
        )


        worksheet.cell(
            row=current_row,
            column=5,
            value=total_lectures
        )


        worksheet.cell(
            row=current_row,
            column=6,
            value=present_lectures
        )


        worksheet.cell(
            row=current_row,
            column=7,
            value=absent_lectures
        )


        percentage_cell = worksheet.cell(
            row=current_row,
            column=8,
            value=attendance_percentage
        )


        percentage_cell.number_format = (
            "0.00%"
        )


        current_row += 1


    # ========================================================
    # 22. COLUMN WIDTHS
    # ========================================================

    column_widths = {

        "A": 15,

        "B": 25,

        "C": 12,

        "D": 12,

        "E": 17,

        "F": 12,

        "G": 12,

        "H": 18

    }


    for column, width in (
        column_widths.items()
    ):

        worksheet.column_dimensions[
            column
        ].width = width


    # ========================================================
    # 23. SAVE EXCEL TO MEMORY
    # ========================================================

    excel_file = BytesIO()


    workbook.save(
        excel_file
    )


    excel_file.seek(0)


    # ========================================================
    # 24. CREATE FILE NAME
    # ========================================================

    safe_class_name = (
        classroom.name
        .replace(" ", "_")
    )

    safe_section = (
        classroom.section
        .replace(" ", "_")
    )

    safe_subject_name = (
        subject.name
        .replace(" ", "_")
    )


    filename = (
        f"attendance_"
        f"{safe_class_name}_"
        f"{safe_section}_"
        f"{safe_subject_name}_"
        f"{period}.xlsx"
    )


    # ========================================================
    # 25. RETURN EXCEL FILE
    # ========================================================

    return StreamingResponse(

        excel_file,

        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument"
            ".spreadsheetml.sheet"
        ),

        headers={

            "Content-Disposition":
                (
                    f'attachment; '
                    f'filename="{filename}"'
                )

        }

    )
    
    # ============================================================
# ADMIN ATTENDANCE REPORT
#
# Returns class + section + subject attendance report
# for display inside the Admin Reports page.
#
# This is intentionally a separate endpoint so the existing
# Excel export endpoint above is not changed.
# ============================================================

@router.get("/admin-report")
def get_admin_attendance_report(

    class_name: str = Query(...),

    section: str = Query(...),

    subject_name: str = Query(...),

    period: str = Query(
        "whole",
        description=(
            "Report period: "
            "weekly, monthly, whole or custom"
        )
    ),

    from_date: date | None = Query(None),

    to_date: date | None = Query(None),

    db: Session = Depends(get_db),

    current_user: User = Depends(
        get_current_user
    )
):

    # ========================================================
    # 1. ADMIN CHECK
    # ========================================================

    if current_user.role != "ADMIN":

        raise HTTPException(
            status_code=403,
            detail="Administrator access required"
        )


    # ========================================================
    # 2. CLEAN INPUT
    # ========================================================

    class_name = class_name.strip()

    section = section.strip().upper()

    subject_name = subject_name.strip()

    period = period.strip().lower()


    # ========================================================
    # 3. VALIDATE PERIOD
    # ========================================================

    allowed_periods = {
        "weekly",
        "monthly",
        "whole",
        "custom"
    }

    if period not in allowed_periods:

        raise HTTPException(
            status_code=400,
            detail=(
                "Invalid report period. "
                "Use weekly, monthly, whole or custom."
            )
        )


    # ========================================================
    # 4. VALIDATE CUSTOM DATE RANGE
    # ========================================================

    if period == "custom":

        if not from_date or not to_date:

            raise HTTPException(
                status_code=400,
                detail=(
                    "From date and To date are required "
                    "for custom report."
                )
            )

        if from_date > to_date:

            raise HTTPException(
                status_code=400,
                detail=(
                    "From date cannot be after To date."
                )
            )


    # ========================================================
    # 5. FIND CLASS
    # ========================================================

    classroom = (
        db.query(Class)
        .filter(
            Class.school_id ==
            current_user.school_id,

            func.lower(Class.name) ==
            class_name.lower(),

            func.upper(Class.section) ==
            section,

            Class.is_active == True
        )
        .first()
    )


    if not classroom:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Class '{class_name}' "
                f"with section '{section}' not found"
            )
        )


    # ========================================================
    # 6. FIND SUBJECT
    # ========================================================

    subject = (
        db.query(Subject)
        .filter(
            Subject.school_id ==
            current_user.school_id,

            func.lower(Subject.name) ==
            subject_name.lower(),

            Subject.is_active == True
        )
        .first()
    )


    if not subject:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Subject '{subject_name}' not found"
            )
        )


    # ========================================================
    # 7. VERIFY SUBJECT IS ASSIGNED TO CLASS
    # ========================================================

    class_subject = (
        db.query(ClassSubject)
        .filter(
            ClassSubject.class_id ==
            classroom.id,

            ClassSubject.subject_id ==
            subject.id,

            ClassSubject.is_active == True
        )
        .first()
    )


    if not class_subject:

        raise HTTPException(
            status_code=404,
            detail=(
                f"Subject '{subject.name}' "
                f"is not assigned to "
                f"{classroom.name} - "
                f"Section {classroom.section}"
            )
        )


    # ========================================================
    # 8. DETERMINE DATE RANGE
    # ========================================================

    today = date.today()

    report_from_date = None

    report_to_date = today


    # --------------------------------------------------------
    # WEEKLY
    # --------------------------------------------------------

    if period == "weekly":

        report_from_date = (
            today -
            timedelta(
                days=today.weekday()
            )
        )


    # --------------------------------------------------------
    # MONTHLY
    # --------------------------------------------------------

    elif period == "monthly":

        report_from_date = today.replace(
            day=1
        )


    # --------------------------------------------------------
    # WHOLE
    # --------------------------------------------------------

    elif period == "whole":

        report_from_date = None

        report_to_date = today


    # --------------------------------------------------------
    # CUSTOM
    # --------------------------------------------------------

    elif period == "custom":

        report_from_date = from_date

        report_to_date = to_date


    # ========================================================
    # 9. GET ACTIVE STUDENTS
    # ========================================================

    students = (
        db.query(Student)
        .filter(
            Student.school_id ==
            current_user.school_id,

            Student.class_id ==
            classroom.id,

            Student.is_active == True
        )
        .order_by(
            Student.roll_number
        )
        .all()
    )


    if not students:

        raise HTTPException(
            status_code=404,
            detail="No active students found"
        )


    # ========================================================
    # 10. ATTENDANCE QUERY
    # ========================================================

    attendance_query = (
        db.query(Attendance)
        .filter(
            Attendance.school_id ==
            current_user.school_id,

            Attendance.class_id ==
            classroom.id,

            Attendance.subject_id ==
            subject.id
        )
    )


    # ========================================================
    # 11. DATE FILTER
    # ========================================================

    if report_from_date is not None:

        attendance_query = (
            attendance_query
            .filter(
                Attendance.attendance_date
                >= report_from_date
            )
        )


    if report_to_date is not None:

        attendance_query = (
            attendance_query
            .filter(
                Attendance.attendance_date
                <= report_to_date
            )
        )


    # ========================================================
    # 12. GET ATTENDANCE
    # ========================================================

    attendance_records = (
        attendance_query
        .order_by(
            Attendance.attendance_date
        )
        .all()
    )


    # ========================================================
    # 13. FIND TOTAL LECTURES
    #
    # One unique attendance date = one lecture.
    # ========================================================

    lecture_dates = {

        attendance.attendance_date

        for attendance
        in attendance_records

    }


    total_lectures = len(
        lecture_dates
    )


    # ========================================================
    # 14. CREATE ATTENDANCE LOOKUP
    # ========================================================

    attendance_map = {}


    for attendance in attendance_records:

        student_id = attendance.student_id

        if student_id not in attendance_map:

            attendance_map[
                student_id
            ] = []

        attendance_map[
            student_id
        ].append(attendance)


    # ========================================================
    # 15. BUILD REPORT
    # ========================================================

    report_rows = []


    for student in students:

        student_records = (
            attendance_map.get(
                student.id,
                []
            )
        )


        # ----------------------------------------------------
        # PRESENT
        # ----------------------------------------------------

        present_count = sum(

            1

            for attendance
            in student_records

            if attendance.status
            and attendance.status.upper()
            in {
                "PRESENT",
                "LATE"
            }

        )


        # ----------------------------------------------------
        # ABSENT
        # ----------------------------------------------------

        absent_count = (
            total_lectures
            - present_count
        )


        if absent_count < 0:

            absent_count = 0


        # ----------------------------------------------------
        # PERCENTAGE
        # ----------------------------------------------------

        if total_lectures > 0:

            attendance_percentage = (
                present_count
                / total_lectures
            ) * 100

        else:

            attendance_percentage = 0


        # ----------------------------------------------------
        # REPORT ROW
        # ----------------------------------------------------

        report_rows.append({

            "roll_number":
                student.roll_number,

            "name":
                student.name,

            "div":
                student.section
                or classroom.section,

            "gender":
                student.gender,

            "total_lectures":
                total_lectures,

            "present":
                present_count,

            "absent":
                absent_count,

            "percentage":
                round(
                    attendance_percentage,
                    2
                )

        })


    # ========================================================
    # 16. RESPONSE
    # ========================================================

    return {

        "class_name":
            classroom.name,

        "section":
            classroom.section,

        "subject_name":
            subject.name,

        "subject_id":
            subject.id,

        "period":
            period,

        "from_date":
            report_from_date,

        "to_date":
            report_to_date,

        "total_lectures":
            total_lectures,

        "student_count":
            len(report_rows),

        "students":
            report_rows

    }